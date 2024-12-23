import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
chrome_options = Options()
import pandas as pd
import time
from openpyxl import load_workbook, Workbook
import os
from logger import init_logger

logger = init_logger()

dir_path=os.path.join(str(os.getcwd()), os.getenv("company_name"))
target_workbook_path=os.path.join(dir_path,'Nth Tier Mapping.xlsx')

def check_and_copy_workbook(existing_workbook_path, sheet_name_to_copy, target_workbook_path=target_workbook_path):
    # Load the existing workbook from which data will be copied
    if not os.path.exists(existing_workbook_path):
        logger.info(f"Error: The source workbook {existing_workbook_path} does not exist.")
        return
 
    source_wb = load_workbook(existing_workbook_path)
    
    source_sheet = source_wb['Sheet1']
    
     # Check if the target workbook exists
    if os.path.exists(target_workbook_path):
        logger.info(f"{target_workbook_path} already exists.")
        wb=load_workbook(target_workbook_path)
        new_sheet= wb.create_sheet(title=sheet_name_to_copy)
    else:
        logger.info(f"{target_workbook_path} not found, creating a new one.")
        wb = Workbook()
        new_sheet= wb.create_sheet(title=sheet_name_to_copy)
 
    # Copy the data from the source sheet to the new sheet
    for row in source_sheet.iter_rows(values_only=True):
        new_sheet.append(row)
 
    # Save the new workbook
    wb.save(target_workbook_path)
    logger.info(f"New workbook '{target_workbook_path}' created and data copied successfully.")

import os
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
chrome_options = Options()
import pandas as pd
import time
from openpyxl import load_workbook, Workbook
from wb_utils import check_and_copy_workbook as update_sheet


from logger import init_logger

logger = init_logger()

def get_companies(company, file_path):

    driver = webdriver.Chrome()
    url="https://www.importgenius.com/search/"+company
    driver.get(url)

    soup = BeautifulSoup(driver.page_source,"html.parser")
    pagination_data = str(soup.find('div', class_ = "company-list-nav my-4").text).split("\n")[1].strip()
    card_qnt = list(map(lambda x:int(x),re.findall(r'\d+',pagination_data)))
    clicker_count = int(card_qnt[2]/(card_qnt[1] - card_qnt[0]))
    company_info={"company_name":[], "company_location":[], "site":[], "relation":[], "check":[]}

    for page in range(clicker_count):

        cards = soup.find_all('div', class_ = "card shadow-sm")

        for card in cards:

            company_name = str(card.find('div',class_="company-name").text).replace("\n","").strip()
            company_location = str(card.find('div',class_="company-location").text).replace("\n","").strip()
            href = card.find("a").attrs.get('href')
            site = 'https://www.importgenius.com'+ href

            company_info.get("company_name").append(company_name)
            company_info.get("company_location").append(company_location)
            company_info.get("site").append(site)
            company_info.get("relation").append('Company')
            company_info.get("check").append('')
            
        
        # breakpoint()
        if clicker_count >1:
            button=driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[2]/div/div[2]/div[1]/div/div/div[2]/div/div[2]/nav/ul/li[8]/a')
            button.click()

    df = pd.DataFrame(company_info)
    df=df.drop_duplicates()
    df.to_excel(file_path, index=False)

    driver.quit()

def update_check(index, site, file_path):
    # file_path = 'company_.xlsx'

    # Load the Excel file into a pandas DataFrame
    df = pd.read_excel(file_path, sheet_name='Sheet1')

    # Check for rows where both index and site match
    mask = df['site'] == site

    # Update the 'check' column for matching rows
    df.loc[mask, 'check'] = 'Done'

    # Save the updated DataFrame back to the Excel file
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df.to_excel(writer, sheet_name='Sheet1', index=False)


def fetch_list(site, index, company, file_path):
    driver = webdriver.Chrome()

    driver.get(site)

    soup = BeautifulSoup(driver.page_source,"html.parser")
    time.sleep(1)
    tabs=soup.find('ul',class_="tabs tab-wrapper-paywall")
    company_info={"company_name":[], "company_location":[], "site":[], "relation":[], "check":[]}
    try:
        if 'Suppliers' in tabs.text:
            driver.find_element(By.XPATH,"//*[@id='company-profile']/div/div[3]/div/nav/ul/li[4]/a").click()
            tables = soup.find_all('table', class_="table text-nowrap")
            table = tables[2].find('tbody')

            
            rows = table.find_all("tr")
            for row in rows:
                cols = row.find_all("td")
                supplier_name=str(cols[1].text).replace("\n","").strip()
                location=str(cols[2].text).replace("\n","").strip()
                href=cols[1].find("a").attrs.get('href')
                site1 = 'https://www.importgenius.com'+ href
                company_info.get("company_name").append(supplier_name)
                company_info.get("company_location").append(location)
                company_info.get("site").append(site1)
                value = f'supplier for {index + 2} {company}'
                company_info.get("relation").append(value)
                company_info.get("check").append('')
                
    except:
        pass

    df = pd.DataFrame(company_info)
    
    # file_path='company_.xlsx'
    book=load_workbook(file_path)

    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)

    update_check(index, site, file_path)

    driver.quit()

def process(company='Analog devices'):
    sheet=company.replace(" ",'')
    file_path=sheet+'suppliers'+'.xlsx'
    get_companies(company, file_path)
    while True:
        df = pd.read_excel(file_path, sheet_name='Sheet1')
        df = df.sort_values(by=['company_name', 'company_location', 'site', 'check'], ascending=[True, True, True, False])
        df = df.drop_duplicates(subset=['company_name', 'company_location', 'site'], keep= 'first')
        filter_df = df[df['check'] != 'Done']
        col_data=filter_df['site'].tolist()
        logger.info(f"{len(filter_df)},processing in loop--------")
        if len(filter_df) <=0:
            break

        for index, row in filter_df.iterrows():
            if row['check'] !='Done':
                site=row['site']
                company=row['company_name']
                fetch_list(site, index, company, file_path)

    update_sheet(existing_workbook_path=file_path, sheet_name_to_copy='IG_suppliers')
    os.unlink(file_path)